# !pip install regex openpyxl  # Uncomment this line if running locally or if Colab needs it
import sys

try:
    import regex as re
except ImportError:
    print("⚠️ The 'regex' module is required for advanced Japanese matching.")
    print("   Please run: pip install regex")
    sys.exit(1)

from google.colab import files
import io
from openpyxl import load_workbook
import os

# ============================================================
#  MODE SELECTION
# ============================================================
print("Choose a mode:\n"
      "  1. Insert delimiters into an Excel file (Processing)\n"
      "  2. Linebreak a Japanese text segment into balanced chunks (Splitting)\n")

mode = input("Enter 1 or 2 (default: 1): ").strip() or "1"

# --- Main regex (Unchanged) ---
pattern = re.compile(r"""
(
    (\p{Han}{1,2}|\p{Katakana}{2,12}|こと|ところ|\p{Han}(?:\p{Hiragana}(?!で))+\p{Han}|\p{Katakana}{2,12}\p{Han}|もの|入り|」|たち|ここ|そこ|\p{Han}ら|(?P<double>\p{Hiragana}{2})(?P=double)|[えけげせぜてでねめれ]る|まま|[あこそ]いつ|あ[なん]た|さん|まみれ|おそらく|たっぷり|気持ち|すら|さすが|くず|あちこち|もと|さま|[こそあど]れ|ど[れん]だけ|みんな|やつ|[あこそ]いつ|すで|だ|[こそあ]ちら|[こそあ]っち)
    (が(?!(して|った|ら))|か(?!([はもらなえがけげせぜてでねめれいきぎしちにんをうくぐすつぬむるりっ]|った|さ))|か[は]|は(?!ず)|も(?!の)|の(?![みにがはた為よ])|なく(?!て)|な(?![くのんらるいし])|する(?!な)|から(?!して)|まで|に(?!([はも]|ついて|関して|すら))|
    に[はも]|へ[の]|へ(?![の])|で(?![はもすしきの])|で[はも]|じて(?!る)|や(?![からりるれ])|と[のはか]|と(?!([のなはかす]|[い言云]う))|して[はも]|して(?![はもる])|ならば|なら(?![ばで]))
    |
    [、。？！・：；]
    |
    (――)|(……)|(\.\.\.)
    |
    \p{Han}すぎ[^たるだ]
    |
    について(?![はも])|について[はも]|に関して(?![はも])|に関して[はも]|[っいきぎしちにん][ただ]り|とにかく|でも|[くぐ]らい(?!は)|[くぐ]らいは|まるで|って(?![るたかも])|っても|
    すなわち|[うくぐすつぬふむる]の[にはもが]|を|んな[のに]|[って]たら|として(?!も)|つまり|ちょっと|ちょうど|々な|々に(?![もは])|々に[もは]|たい(?=\p{Han})|けど|よう[なに](?=(\p{Han}{2}|\p{Katakana}{2}))|
    だと(?!は)|だとは|とは|[のただ]ほうが|ないほうが|[のただ]方が|ない方が|風に|[いきしちにひみり]たくて|[うくすつぬふむる]まて|[^一-龯]続く|ないと(?=いけ)|く(?=(\p{Han}|\p{Katakana}{2}))|
    ほとんど|らしくて(?!は)|らしく(?!て)|ため([にの](?![はも])|ならば|なら(?!ば))|ため[にの][はも]|為に(?![はも])|為に[はも]|わけ(では|じゃ(?!あ))|ほうが(?=(\p{Han}|\p{Katakana}{2}))|
    いきなり|すれば|(れば|ないと)(?=([い良善好]い|[よ良善好]か))|て(?=い?ました)|しっかり|して(?=あげ([るた]|(ます|まし)))|て(?=(ください|下さい|ちょうだい))|これまでに(?!は)|
    より(?=ずっと)|はじめて|[てで](?=くれ)|くなって(?!は)|され[るた](?![んの])|かった(?![んのりわっがぞぜ])|もなくて(?!は)|あらゆる|すべて(の|を|では|じゃ(?!あ))|すぐに[はも]|すぐに(?![はも])|
    もなく(?!て)|ながら|がてら|った(?![らんのりわっがぞぜ])|よりも|かも(?=[しれ])|とともに(?![はも])|と共に(?![はも])|もっとも|すべて(?!でのを)|ただの|まま(?=(\p{Han}|\p{Katakana}{2}))|
    どうして|どうやって|した(?=(\p{Han}{2}|こと|とこ))|のもとに|[うくすつぬふむるじの]よう[にな]|れて(?=(いき?ま|いる|いた|いな))|じゃ(?=な[いか])|では(?=な[いか])|またしても|
    どうなるか(?!は)|どうなるかは|しばらく|[えけげせぜてでねめれ]なく(?!て)|[えけげせぜてでねめれあかさたなまら]ずに|[えけげせぜてでねめれいきしじちにみりっ]て(?=い(る|ま|く|け))|
    \p{Han}し?い(?=(\p{Han}|\p{Katakana}{2}))(?!出)|\p{Han}しく(?=(\p{Han}|\p{Katakana}{2}))|べきじゃ(?!あ)|かなり(?=(\p{Han}|\p{Katakana}{2}))|[えけげせぜてでねめれ]ば(?=(\p{Han}|\p{Katakana}{2}))|
    ゆっくり(?=(\p{Han}|\p{Katakana}{2}))|ちゃんと(?=(\p{Han}|\p{Katakana}{2}))|(なければ|なきゃ|ないと)(?=(なら|いけ))|\p{Hiragana}(?=(はず|べき)(だ|よ|$|。|…|！|？))|\p{Hiragana}(?=\p{Katakana}{2})|て(?=ありがと)|
    なら(?=(\p{Han}|\p{Katakana}{2}))|なのは|[えけげせぜてでねめれ][るてた](?=(\p{Han}|\p{Katakana}{2}))|たく(?=な[いか])|[わかさたなまら]れ[るた](?=(\p{Han}|\p{Katakana}{2}))|いくつか|\p{Han}ても|して(?=(\p{Han}|\p{Katakana}{2}))|
    \p{Han}たる(?=(\p{Han}|\p{Katakana}{2}))|という(?=(\p{Han}|\p{Katakana}{2}))|を|な[くい](?=(\p{Han}|\p{Katakana}{2}))|\p{Han}\p{Hiragana}に(?=な(る|った|らな))|いた(?=(\p{Han}|\p{Katakana}{2}))|
    ないと(?=(\p{Han}|\p{Katakana}{2}))|て(?=ほし[いくか])|\p{Han}{2}(?=\p{Katakana}{2})|な(?=(\p{Han}|\p{Katakana}{2}))|\p{Katakana}{2}(?=\p{Han}{2})|(?P<doubler>\p{Hiragana}{2})(?P=doubler)|くて(?=\p{Han})|
    しか(?=(\p{Han}|\p{Katakana}{2}))|よりかは|て(?=しま[ういわ])|とっ?ても|\p{Han}\p{Hiragana}(?=\p{Han}{2})|とか(?=\p{Han})|もう(?=\p{Han})|\p{Hiragana}(?=つもり)|が(?=(\p{Han}{2}|\p{Katakana}{2}))
)
""", re.VERBOSE)

# ============================================================
#  MODE 1: Excel delimiter insertion
# ============================================================
if mode == "1":
    user_input = input("Enter a delimiter (press Enter for invisible ZWSP '\\u200B'): ").strip()
    INSERT_CHAR = user_input if user_input else '\u200B'
    preview_symbol = "[ZWSP]" if INSERT_CHAR == '\u200B' else INSERT_CHAR
    print(f"✅ Using delimiter: {repr(INSERT_CHAR)}")
    print(f"🔍 Preview: 日本語{preview_symbol}テキスト")

    print("\n📂 Please upload your Excel file:")
    uploaded = files.upload()
    
    if not uploaded:
        print("⚠️ No file uploaded. Exiting.")
    else:
        filename = list(uploaded.keys())[0]
        wb = load_workbook(io.BytesIO(uploaded[filename]))
        target_headers = {"ja", "jp", "jap", "japanese"}

        def postprocess_ellipses(text):
            if not isinstance(text, str): return text
            # Fix ellipses that might have been split awkwardly
            text = re.sub(rf'^(…{{1,4}}){re.escape(INSERT_CHAR)}', r'\1', text)
            text = re.sub(r'(?<!…)(…)(?!…)(?=\S)', lambda m: m.group(1) + INSERT_CHAR, text)
            text = re.sub(rf'([^\s…]){re.escape(INSERT_CHAR)}(…|\.\.\.)', r'\1\2', text)
            return text

        def insert_delimiter(text):
            if not isinstance(text, str): return text
            
            def replacer(m):
                end = m.end()
                remainder = text[end:]
                
                # Safety check for end of string
                if not remainder:
                    return m.group(0)

                next_char = remainder[0]
                # Logic: Do NOT insert delimiter if the next character is punctuation
                # or if the rest of the cell is only punctuation/whitespace.
                if re.match(r'[、。？！,．,.!?"”」』）)]', next_char) or re.match(r'^[、。？！…‥！？\s]*$', remainder):
                    return m.group(0)
                
                return m.group(0) + INSERT_CHAR

            processed = pattern.sub(replacer, text)
            return postprocess_ellipses(processed)

        print("⏳ Processing...")
        processed_count = 0
        
        for ws in wb.worksheets:
            # Create header mapping (Header Name -> Column Index)
            headers = {}
            for cell in ws[1]:
                if cell.value:
                    headers[cell.value] = cell.column

            for header, col in headers.items():
                if str(header).strip().lower() in target_headers:
                    for row in range(2, ws.max_row + 1):
                        cell = ws.cell(row=row, column=col)
                        if cell.value and isinstance(cell.value, str):
                            new_val = insert_delimiter(cell.value)
                            if new_val != cell.value:
                                cell.value = new_val
                                processed_count += 1

        name, ext = os.path.splitext(filename)
        output_filename = f"delimiters_added_{name}{ext}"
        wb.save(output_filename)
        files.download(output_filename)
        print(f"✅ Done! Processed {processed_count} cells.")
        print(f"⬇️ File saved as: {output_filename}")

# ============================================================
#  MODE 2: Smart text segment linebreaker
# ============================================================
elif mode == "2":
    text = input("Paste the Japanese text segment:\n").strip()
    try:
        lines_input = input("How many lines would you like to split it into? ").strip()
        lines = int(lines_input)
        if lines < 1: raise ValueError
    except ValueError:
        print("⚠️ Invalid line count. Defaulting to 2.")
        lines = 2

    # Find all potential breakpoints
    break_positions = [m.end() for m in pattern.finditer(text)]
    
    if not break_positions:
        print("⚠️ No suitable breakpoints found in the text.")
        print(f"Original: {text}")
    else:
        total_len = len(text)
        target_len = total_len / lines
        chosen_breaks = []
        last = 0

        # Select the best breakpoints closest to mathematical division
        for i in range(1, lines):
            target_pos = target_len * i
            
            # Filter valid breaks that are ahead of the last one
            valid_breaks = [b for b in break_positions if b > last]
            
            if not valid_breaks:
                break # No more breaks available
            
            best_break = min(valid_breaks, key=lambda x: abs(x - target_pos))
            chosen_breaks.append(best_break)
            last = best_break

        chosen_breaks = sorted(set(chosen_breaks))
        
        # Construct chunks
        chunks = []
        prev = 0
        for bp in chosen_breaks:
            chunks.append(text[prev:bp])
            prev = bp
        chunks.append(text[prev:])

        # --- Polishing pass: punctuation fixes (Kinsoku Shori) ---
        def polish_lines(chunks):
            adjusted = chunks[:]
            punct_start = "、。？！：；…‥" + "..."
            
            # RULE 1: Fix "orphaned" leading punctuation.
            # If a line starts with punctuation, move it to the end of the previous line.
            for i in range(1, len(adjusted)):
                # Check for leading punctuation (1 to 3 chars length to catch "...")
                m = re.match(rf'^([{re.escape(punct_start)}]{{1,3}})', adjusted[i])
                if m:
                    tok = m.group(1)
                    # Attach to previous line
                    adjusted[i-1] += tok
                    # Remove from current line
                    adjusted[i] = adjusted[i][len(tok):]

            # Cleanup: remove empty lines created by shifting
            final = [c for c in adjusted if c]
            return final

        chunks = polish_lines(chunks)

        print("\n✅ Suggested linebreaks:\n")
        for i, chunk in enumerate(chunks, 1):
            print(f"{i:02d}: {chunk}")

else:
    print("⚠️ Invalid mode selected.")
